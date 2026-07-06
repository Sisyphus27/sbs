"""One-time import: cold-backup xlsx -> Profile. Specific to the user's 4x layout."""
import openpyxl
from .data.schema import Lift, Profile

QS_MAIN_ROWS = [5, 6, 7, 8]            # Squat/Bench/DL/OHP
QS_AUX_ROWS = [11, 12, 13, 14, 15, 16]  # Front/Paused/Close Grip/Long Pause/RDL/Incline

# SBS-tier defaults the user can edit afterwards.
SBS_DEFAULTS = {
    5:  (0.75, 4, 8, 3), 6: (0.75, 4, 8, 3), 7: (0.80, 3, 6, 3), 8: (0.75, 4, 8, 3),
    11: (0.75, 4, 8, 3), 12: (0.75, 4, 8, 3), 13: (0.75, 4, 8, 3),
    14: (0.75, 4, 8, 3), 15: (0.75, 4, 8, 3), 16: (0.75, 4, 8, 3),
}


def _is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


def import_profile(xlsx_path: str, sheet: str = "4x") -> Profile:
    wb = openpyxl.load_workbook(xlsx_path)
    qs = wb["Quick Setup"]
    lifts = []

    qs_rows = list(QS_MAIN_ROWS) + list(QS_AUX_ROWS)
    for r in qs_rows:
        name = qs[f"C{r}"].value
        one_rm = qs[f"D{r}"].value
        if not name or one_rm is None:
            continue
        intensity, reps, repout, sets = SBS_DEFAULTS[r]
        # main vs aux drives the schedule ladder the engine reads (Task 4):
        # main lifts follow MAIN_LADDER, aux follow AUX_LADDER.
        kind = "main" if r in QS_MAIN_ROWS else "aux"
        lifts.append(Lift(name=str(name), tier="sbs", day=0, max=float(one_rm),
                          intensity=intensity, reps=reps, repout=repout, sets=sets,
                          lift_kind=kind))

    ws = wb[sheet]
    acc_rows = [r for r in range(1, ws.max_row + 1)
                if ws.cell(row=r, column=1).value == "Accessories"]
    days_per_week = len(acc_rows)

    # back rows = non-formula name + numeric weight, immediately above each Accessories label
    for day_idx, label_row in enumerate(acc_rows, start=1):
        r = label_row - 1
        while r > 0:
            a = ws.cell(row=r, column=1).value
            b = ws.cell(row=r, column=2).value
            if a is None and b is None:
                r -= 1
                continue
            if isinstance(a, str) and not _is_formula(a) and isinstance(b, (int, float)):
                lifts.append(Lift(name=a, tier="t2", day=day_idx, start=float(b)))
                r -= 1
                continue
            break  # hit SBS formula rows or a Day label -> stop scanning up

    # accessories = non-formula name + numeric weight, in rows below each Accessories label
    for day_idx, label_row in enumerate(acc_rows, start=1):
        next_boundary = (acc_rows[day_idx] if day_idx < len(acc_rows) else ws.max_row + 1)
        for r in range(label_row + 1, next_boundary):
            a = ws.cell(row=r, column=1).value
            b = ws.cell(row=r, column=2).value
            if a is None and b is None:
                break   # end of this day's contiguous accessory block
            if isinstance(a, str) and not _is_formula(a) and isinstance(b, (int, float)):
                lifts.append(Lift(name=a, tier="t3", day=day_idx, start=float(b)))

    # SBS lifts day assignment: distribute round-robin by order (user edits after)
    sbs_lifts = [l for l in lifts if l.tier == "sbs"]
    for i, l in enumerate(sbs_lifts):
        l.day = (i % days_per_week) + 1

    return Profile(rounding=2.5, days_per_week=days_per_week, incr=2.5,
                   t2_reset_pct=0.70, t2_fail=3, t3_target=15, lifts=lifts)
