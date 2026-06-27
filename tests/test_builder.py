import openpyxl, shutil, os
import pytest
from tools.sbs_gzclp import builder

SRC = r"D:\WorkSpace\sbs\SBS RTF filled GZCLP.xlsx"

def _copy():
    dst = os.path.join(os.path.dirname(SRC), "_test_work.xlsx")
    shutil.copy(SRC, dst)
    return dst

def test_inject_config_writes_params_and_slot_tables():
    dst = _copy()
    try:
        builder.inject_config(dst)
        wb = openpyxl.load_workbook(dst)
        ws = wb["Quick Setup"]
        assert ws["A70"].value == "T2/T3 Parameters"
        assert ws["B71"].value == 2.5      # T2_incr
        assert ws["B72"].value == 0.8      # T2_reset
        assert ws["B73"].value == 3        # T2_fail
        assert ws["B74"].value == 15       # T3_target
        assert ws["B75"].value == 2.5      # T3_incr
        assert ws["A77"].value == "Back (T2) slots"
        assert ws["A93"].value == "Accessories (T3) slots"
        assert isinstance(ws["A79"].value, str)          # default back lift name
        assert isinstance(ws["B79"].value, (int, float)) # start weight
        # Preservation: an existing SBS formula in the 2x sheet must survive.
        assert str(wb["2x"]["B5"].value).startswith("=MROUND(B4*Setup!B3")
        # Second T2 slot row (Pull-ups) - guards against off-by-one.
        assert ws["A80"].value == "Pull-ups"
        assert ws["B80"].value == 20.0
        assert ws["C79"].value == 1  # Day column for slot 1
        # High slots are now seeded (no collision / empty zero cells).
        assert isinstance(ws["A90"].value, str)                   # last T2 slot
        assert isinstance(ws["B111"].value, (int, float))         # last T3 slot weight
    finally:
        os.remove(dst)


from tools.sbs_gzclp.columns import BLOCKS, col

# Sheets that should receive an Accessories zone under the standard template.
DAY_SHEETS = ["2x", "3x", "4x", "5x", "6x"]


def _assert_t3_zone_shape(wb, sheet_name):
    """Core invariants for one day sheet's T3 zone after inject_t3_zones."""
    ws = wb[sheet_name]
    acc_rows = [r for r in range(1, ws.max_row + 1)
                if ws.cell(row=r, column=1).value == "Accessories"]
    assert acc_rows, f"no Accessories label found on {sheet_name}"
    first_data = acc_rows[0] + 1

    # week-1 weight cell (B) references the Quick Setup T3 start weight
    b = ws[f"B{first_data}"].value
    assert isinstance(b, str) and b.startswith("='Quick Setup'"), (
        f"{sheet_name}!B{first_data} should be a Quick Setup T3 ref, got {b!r}"
    )
    # third accessory row is also populated (3rd lift per day, k=2)
    b3 = ws[f"B{first_data + 2}"].value
    assert (isinstance(b3, str) and b3.startswith("='Quick Setup'")), (
        f"{sheet_name}!B{first_data + 2} should be a Quick Setup T3 ref, got {b3!r}"
    )
    # week-2 weight cell (I) is the T3 progression formula
    i = ws[f"I{first_data}"].value
    assert isinstance(i, str) and i.startswith("=IF(F"), (
        f"{sheet_name}!I{first_data} should start with =IF(F..., got {i!r}"
    )
    assert "MROUND(B" in i and "Quick Setup" in i
    # target cell (C) references T3_target param
    c = ws[f"C{first_data}"].value
    assert c == "='Quick Setup'!$B$74", f"{sheet_name}!C{first_data} = {c!r}"
    # Last block (BLOCKS[-1] = "EL") must carry a progression formula on the
    # first T3 data row — guards against the loop omitting the final week.
    last_block = BLOCKS[-1]
    el = ws[f"{last_block}{first_data}"].value
    assert isinstance(el, str) and el.startswith("=IF("), (
        f"{sheet_name}!{last_block}{first_data} should start =IF(...), got {el!r}"
    )
    # SBS untouched: a known SBS cell still intact
    assert str(wb[sheet_name]["B5"].value).startswith("=MROUND(B4*Setup!B3")


def test_inject_t3_zones_writes_progression_formulas():
    dst = _copy()
    try:
        builder.inject_config(dst)
        builder.inject_t3_zones(dst)
        wb = openpyxl.load_workbook(dst)
        # Loop core checks across every day sheet, not just 2x — a 4x/5x/6x-only
        # regression (e.g. wrong max_row handling) would otherwise slip through.
        for sh in DAY_SHEETS:
            _assert_t3_zone_shape(wb, sh)
    finally:
        os.remove(dst)


def test_inject_t3_zones_raises_on_missing_room():
    """A template whose Accessories section is too short must fail loudly."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2x"
    ws["A1"] = "Accessories"   # label at row 1
    ws["A2"] = "only one row below"
    # max_row is 2 -> only 1 row below the label, but ACCESSORIES_PER_DAY=3
    # are required. Validation must raise before any writes happen.
    tmp = "_t3_validation.xlsx"
    wb.save(tmp)
    try:
        with pytest.raises(ValueError, match="Accessories label at row 1"):
            builder.inject_t3_zones(tmp)
    finally:
        os.remove(tmp)


BACK_PER_DAY = 2


def test_inject_t2_zones_writes_state_machine_rows():
    dst = _copy()
    try:
        builder.inject_config(dst)
        builder.inject_t2_zones(dst)
        wb = openpyxl.load_workbook(dst)
        ws = wb["2x"]
        acc_rows = [r for r in range(1, ws.max_row+1)
                    if ws.cell(row=r, column=1).value == "Accessories"]
        first_back_state = acc_rows[0] - BACK_PER_DAY * 2
        # week-1 seed
        assert ws[f"B{first_back_state}"].value == "='Quick Setup'!B79"   # T2 slot 1 start (row 79)
        assert ws[f"C{first_back_state}"].value == 10
        assert ws[f"D{first_back_state}"].value == 0
        # week-2 state formulas (I/J/K block)
        assert ws[f"J{first_back_state}"].value.startswith("=IF(F")     # target
        assert ws[f"K{first_back_state}"].value.startswith("=IF(F")     # streak
        assert ws[f"I{first_back_state}"].value.startswith("=IF(F")     # weight
        # log row week-1 last-set (F) is blank
        log_row = first_back_state + 1
        assert ws[f"F{log_row}"].value is None
        # SBS untouched
        assert str(wb["2x"]["B5"].value).startswith("=MROUND(B4*Setup!B3")
    finally:
        os.remove(dst)


def test_inject_t2_zones_preserves_day2_formula_refs():
    """insert_rows must translate same-sheet refs so Day2+ SBS formulas still point at their own (shifted) log/TM rows."""
    dst = _copy()
    try:
        builder.inject_config(dst)
        builder.inject_t2_zones(dst)
        wb = openpyxl.load_workbook(dst)
        ws = wb["2x"]
        # Find a Day2 SBS lift: a row whose col A is a CONCATENATE(...," TM") formula below the first Accessories label.
        acc = [r for r in range(1, ws.max_row+1) if ws.cell(row=r, column=1).value == "Accessories"]
        # Day2 is AFTER first Accessories. Find first " TM" row with row > acc[0]:
        day2_tm_rows = [r for r in range(acc[0]+1, ws.max_row+1)
                        if isinstance(ws.cell(row=r, column=1).value, str)
                        and ws.cell(row=r, column=1).value.startswith("=CONCATENATE")]
        assert day2_tm_rows, "expected a Day2 TM row after first Accessories"
        tm = day2_tm_rows[0]
        name_f = ws.cell(row=tm, column=1).value          # =CONCATENATE(A{tm+1}," TM")
        tm_w2  = ws.cell(row=tm, column=9).value          # col I, week-2 TM formula
        # The CONCATENATE must reference THIS lift's own working-weight row (tm+1), not a stale row.
        assert f"A{tm+1}" in name_f, f"name ref broken: {name_f} (expected A{tm+1})"
        # The week-2 TM formula must reference THIS lift's own log row F{tm+1} and own TM K{tm}, not old rows.
        assert f"F{tm+1}" in tm_w2, f"log ref broken: {tm_w2} (expected F{tm+1})"
        assert f"K{tm}" in tm_w2, f"self TM ref broken: {tm_w2} (expected K{tm})"
        # Other-sheet refs untouched:
        assert "Setup!" in tm_w2 and "Quick Setup" in str(ws.cell(row=tm+1,column=9).value)
    finally:
        os.remove(dst)


def test_build_all_runs_in_order_and_all_sheets_have_both_zones():
    dst = _copy()
    try:
        builder.build_all(dst)
        wb = openpyxl.load_workbook(dst)
        for sh in ["2x", "3x", "4x", "5x", "6x"]:
            ws = wb[sh]
            acc = [r for r in range(1, ws.max_row + 1)
                   if ws.cell(row=r, column=1).value == "Accessories"]
            assert len(acc) >= 2, f"{sh} missing day blocks"
            for label in acc:
                # T2 zone: BACK_PER_DAY*2 rows immediately above the label
                first_back = label - BACK_PER_DAY * 2
                assert ws[f"C{first_back}"].value == 10          # T2 seed target
                assert ws[f"I{first_back}"].value.startswith("=IF(F")  # T2 progression
                # T3 zone: row immediately below the label
                t3 = label + 1
                assert str(ws[f"B{t3}"].value).startswith("='Quick Setup'")  # T3 start ref
                assert ws[f"I{t3}"].value.startswith("=IF(F")                # T3 progression
        # SBS T1/aux untouched: a known SBS cell still intact
        assert str(wb["2x"]["B5"].value).startswith("=MROUND(B4*Setup!B3")
        # Quick Setup config present
        qs = wb["Quick Setup"]
        assert qs["B74"].value == 15   # T3_target
        assert qs["A70"].value == "T2/T3 Parameters"
    finally:
        os.remove(dst)


def test_build_all_preserves_sbs_lift_identity():
    """SBS TM cells must keep their 'Quick Setup'!D{X} cross-sheet refs after build_all
    (only same-sheet row refs shift). Compare lift identity (QS D-ref) cold vs built."""
    import re
    dst = _copy()
    try:
        builder.build_all(dst)
        cold = openpyxl.load_workbook(SRC)              # SRC is the pristine real file
        built = openpyxl.load_workbook(dst)
        qsre = re.compile(r"'Quick Setup'!\$?D(\d+)")
        for sh in ["2x","3x","4x","5x","6x"]:
            def ids(ws):
                out=[]
                for r in range(1, ws.max_row+1):
                    v=ws.cell(row=r,column=2).value
                    if isinstance(v,str) and v.startswith("=IF(D") and "'Quick Setup'!D" in v:
                        m=qsre.search(v)
                        if m: out.append(int(m.group(1)))
                return sorted(out)
            assert ids(cold[sh]) == ids(built[sh]), \
                f"{sh} SBS lift identity changed: cold={ids(cold[sh])} built={ids(built[sh])}"
    finally:
        os.remove(dst)
